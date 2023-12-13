
python3 -mzeep "https://saas.hrzucchetti.it/hrpsnaitech-test/servlet/SQLDataProviderServer/SERVLET/a_gv_retry_date?wsdl"

#========================================================================#
# file_collector.py                                                      #
#========================================================================#
#========================================================================#
# usage: file_collector.py [-h] [-f FILE] [-o OUTPUT_FOLDER]             #
#========================================================================#
# optional arguments:
#  -h, --help            show this help message and exit
#  -f FILE, --file FILE  the xlsx file that contains all the files to be
#                        collected
#  -o OUTPUT_FOLDER, --output OUTPUT_FOLDER
#                        the folder in which the collected files are stored
#
#========================================================================#
# Sample usage:                                                          #
#========================================================================#
# 1. Move to the directory where the script is placed.
#   cd "path/to/folder"
# 2. To collect the files listed in the excel file "example.xlsx" and 
# save them into the folder "output"
#   python file_collector.py -f "path/to/sample.xlsx" -o "output"
#========================================================================#

from openpyxl import *
from zeep import *
import os
import argparse

out_path = "images/"
res_path = "results/"
wsdl = "path/to/wsdl"
client = Client(wsdl)
base64_encoded_images = {}
username = "username"
password = "password"
doc_id = "doc_id"

def decode(base64_string, doc, file_id, img_format, out_folder):
    doc_type = ''.join(e for e in doc if e.isalnum())
    try:
        # Attempt to convert the string
        image_64_decode = base64_string
        out_dir = out_path + out_folder + "/"
        if not os.path.exists(out_dir):
            os.makedirs(out_dir)

        # Create a writable image file
        image_result = open(out_dir + doc_type + "_" + file_id + "." + img_format, 'wb')

        # Write the decoded result
        image_result.write(image_64_decode)

        # Close the file after writing image
        image_result.close()
        return image_64_decode
    except:
        print('exception occured in decoding image from ' + str(base64_string))
        raise


def read_images(worksheet):
    wb = load_workbook(worksheet)
    ws = wb['Export Worksheet']
    images = {}
    for c in ws.iter_rows(row_offset=1):
        if c[0].value == None:
            return images
        file_id = c[0].value
        doc_type = c[1].value
        doc_date = c[2].value
        doc_id = c[3].value
        images[doc_id] = [file_id, doc_type, doc_date, doc_id]
        # print([file_id, doc_type, doc_date, doc_id])
    return images


def read_doc_ids(worksheet):
    wb = load_workbook(worksheet)
    ws = wb['Export Worksheet']
    doc_ids = []
    for c in ws.iter_rows(row_offset=1):
        if c[3].value == None:
            return doc_ids
        doc_ids.append(c[3].value)
    return doc_ids


def send_request(doc_id, cl):
    r = cl.service.GetDocumentWithID(document={'username': username,
                                               'password': password,
                                               'doc_id': doc_id})
    return r


def send_raw_request(doc_id, cl):
    with cl.options(raw_response=True):
        r = cl.service.GetDocumentWithID(document={'username': username,
                                                   'password': password,
                                                   'doc_id': doc_id})
    return r


def printProgressBar(iteration, total, prefix='', suffix='', decimals=1, length=100, fill='█'):
    percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
    filled_length = int(length * iteration // total)
    bar = fill * filled_length + '-' * (length - filled_length)
    print('\r%s |%s| %s%% %s' % (prefix, bar, percent, suffix), end='\r')
    # Print New Line on Complete
    if iteration == total:
        print()


def show_license():
    print("\n# ======================================================================================= #\n"
          "# File Collector - Collects image files from the server via SOAP requests                 #\n" +
          "# Copyright © 2018 - Berk Kaan Kuguoglu                                                   #\n" +
          "# ======================================================================================= #\n" +
          "# This program comes with absolutely no warranty.                                         #\n" +
          "# This is free software, and you are welcome to redistribute it under certain conditions. #\n" +
          "# ======================================================================================= #\n" +
          "# usage: python file_collector.py [-h] [-f FILE] [-o OUTPUT_FOLDER]                       #\n" +
          "# ======================================================================================= #\n" +
          "# optional arguments:                                                                     #\n" +
          "# -h, --help show this help message and exit                                              #\n" +
          "# -f FILE, --file the xlsx file that contains all the files to be collected               #\n" +
          "# -o OUTPUT_FOLDER, --output the folder in which the collected files are stored           #\n" +
          "# ======================================================================================= #\n")


def main(args):
    xlsx_file = args.input_sheet
    out_folder = args.output_folder
    # print(xlsx_file)
    # print(out_folder)
    # loop over the images to find the template in
    images = read_images(xlsx_file)
    doc_ids = read_doc_ids(xlsx_file)
    image_ids = []
    num_of_files = len(doc_ids)
    i = 1
    for doc_id in doc_ids:
        percentage = 100 * float(i) / float(num_of_files)
        # print(">>> Collecting files... " +
        #       "[" + str(i) + "/" + str(num_of_files) + "] --- " +
        #       format(percentage, '.2f') + "%")
        prefix = ">>> Collecting files... " + \
                 "[" + str(i) + "/" + str(num_of_files) + "]"
        printProgressBar(i, num_of_files, prefix=prefix + ' Progress:', suffix='Complete', length=25)
        r = send_request(doc_id, client)
        img_format = r['mimeType']
        extension = img_format.split('/')[-1]
        if extension in {"jpeg", "jpg", "png", "JPEG", "JPG"}:
            encoded_img = r['data']
            base64_encoded_images[doc_id] = [encoded_img, extension]
            image_ids.append(doc_id)
        i += 1

    i = 1
    num_of_images = len(image_ids)
    for doc_id in image_ids:
        prefix = ">>> Writing files... [" + str(i) + "/" + str(num_of_images) + "]"
        printProgressBar(i, num_of_images, prefix=prefix + ' Progress:', suffix='Complete', length=25)
        img = base64_encoded_images.get(doc_id)[0]
        img_format = base64_encoded_images.get(doc_id)[1]
        doc_type = images.get(doc_id)[1]
        file_id = images.get(doc_id)[3]
        decode(img, doc_type, file_id, img_format, out_folder)
        i += 1

    return


if __name__ == "__main__":
    # execute only if run as a script
    parser = argparse.ArgumentParser()

    # Add more options if you like
    parser.add_argument("-f", "--file", dest="input_sheet",
                        help="the xlsx file that contains all the files to be collected", metavar="FILE")
    parser.add_argument("-o", "--output", dest="output_folder",
                        help="the folder in which the collected files are stored", metavar="OUTPUT_FOLDER")

    args = parser.parse_args()
    show_license()
    main(args)
