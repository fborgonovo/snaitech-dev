#!C:\Program Files (x86)\Python37-32\python.exe

"""
Read Employee data
"""

import sys
#from openpyxl import load_workbook, Workbook
import openpyxl

if __name__ == "__main__":
    if len(sys.argv) != 2:
    	exit("Need excel file full name.")

employeeFN = sys.argv[1]

employeeWB = openpyxl.load_workbook(employeeFN)
employeeWB_grid_resultSH = employeeWB.active

max_col = employeeWB_grid_resultSH.max_column 

# Loop will print all columns name 
for i in range(1, max_col + 1): 
    cell_obj = employeeWB_grid_resultSH.cell(row = 1, column = i) 
    print(cell_obj.value) 

"""
workbook = Workbook()

sheet_Grid_Result = workbook.active

workbook.save(filename=employeeFN)
"""